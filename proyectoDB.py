from msilib.schema import Font
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pyodbc
import pandas as pd
import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import Workbook


mes = datetime.datetime.now().strftime("%B")
nombreArchivo = 'Inventario de licenciamiento de Grabacion Bogotá, Medellin y Perú (' + \
    mes + '-2022).xlsx'
excelExtensiones = 'Extensiones ('+mes+'-2022).xlsx'
nombreExcel = 'Compilado ('+mes+'-2022).xlsx'  # nombre excel tabla dinamica
borde_doble = Side(border_style="medium")  # Cambiamos el formato del borde
borde_cuadrado = Border(top=borde_doble,
                        right=borde_doble,
                        bottom=borde_doble,
                        left=borde_doble)
centrado = Alignment(horizontal="center")  # Cambiamos el aliniamiento
colorTabla = 'f44336'

# ---------------------------- Archivo Excel


def crearExcelExt():
    wb = openpyxl.Workbook()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Engage - Medellin"
    ws1.append(('Servidor', 'Extension', 'Cliente'))

    ws2 = wb.create_sheet(title="Engage - Peru")
    ws2.append(('Servidor', 'Extension', 'Cliente'))

    ws3 = wb.create_sheet(title="Engage - Bogota")
    ws3.append(('Servidor', 'Extension', 'Cliente'))

    ws4 = wb.create_sheet(title="Uptvity Bogota")
    ws4.append(('Servidor', 'Extension', 'Cliente'))

    ws5 = wb.create_sheet(title="Uptvity Medellin")
    ws5.append(('Servidor', 'Extension', 'Cliente'))

    ws6 = wb.create_sheet(title="Uptivity Despegar")
    ws6.append(('Servidor', 'Extension', 'Cliente'))

    ws7 = wb.create_sheet(title="Ext sin nombre y prueba")
    ws7.append(('Servidor', 'Extension', 'Cliente'))

    wb.save(filename=excelExtensiones)
    wb.close()


def crearExcelFinal():
    wb = openpyxl.Workbook()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Servidores Nice Engage"
    ws2 = wb.create_sheet(title="Servidores Nice Uptivity")
    ws3 = wb.create_sheet(title="Medellin")
    ws3.append(('Cliente','Engage', 'Uptivity','All'))
    wb.save(filename=nombreArchivo)
    wb.close()


#---------------------------- Query
def consultaDbMedEng(cursor):
    cursor.execute("SELECT * FROM Medellin")  # ira el query
    rows = cursor.fetchall()
    return [rows, 1]


def consultaDbBogEng(cursor):
    cursor.execute("SELECT * FROM Bogota")  # ira el query
    rows = cursor.fetchall()
    return [rows, 2]


def consultaDbBogUp(cursor):
    cursor.execute("SELECT * FROM Bog")  # ira el query
    rows = cursor.fetchall()
    return [rows, 3]


def consultaDbMedUp(cursor):
    cursor.execute("SELECT * FROM Med")  # ira el query
    rows = cursor.fetchall()
    return [rows, 4]


def consultaDbDesUp(cursor):
    cursor.execute("SELECT * FROM Des")  # ira el query
    rows = cursor.fetchall()
    return [rows, 5]


#---------------------------- Engage
def guardarExcelEng(consulta):
    wb = load_workbook(excelExtensiones)
    ws1 = wb['Engage - Medellin']
    ws2 = wb['Engage - Peru']
    ws3 = wb['Engage - Bogota']
    ws7 = wb['Ext sin nombre y prueba']
    if consulta[1] == 1:
        for row in consulta[0]:
            a = row[0]
            b = row[1]
            c = row[2]
            if c and c != "Prueba" and c != "NULL":
                if (c != "Bimbo") and (c != "Bimbo Col") and (c != "Bimbo Ecuador") and (c != "CBC Peru") and (c != "Cencosud") and (c != "Crehana") and (c != "Equifax") and (c != "Global Pass ") and (c != "Ricolino Mx") and (c != "Rimac") and (c != "SabMiller Peru") and (c != "SAC Backus"):  # validar los de Peru cuidado con el espacio
                    d = a[row[0].find("AIR"):len(row[0])].replace(" ", "")
                    e = d.replace("0", "")
                    f = c.replace(" ", "")
                    ws1.append([e, b, f])
                else:
                    d = a[row[0].find("AIR"):len(row[0])].replace(" ", "")
                    e = d.replace("0", "")
                    f = c.replace(" ", "")
                    ws2.append([e, b, f])
            else:
                ws7.append(list(row))

    if consulta[1] == 2:
        for row in consulta[0]:
            a = row[0]
            b = row[1]
            c = row[2]
            if c and c != "Prueba" and c != "NULL":
                d = a[row[0].find("AIR"):len(row[0])].replace(" ", "")
                e = d.replace("0", "")
                f = c.replace(" ", "")
                ws3.append([e, b, f])
            else:
                ws7.append(list(row))
    wb.save(filename=excelExtensiones)
    wb.close()


#---------------------------- Uptivity
def guardarExcelUp(consulta):
    wb = load_workbook(excelExtensiones)
    ws4 = wb['Uptvity Bogota']
    ws5 = wb['Uptvity Medellin']
    ws6 = wb['Uptivity Despegar']
    ws7 = wb['Ext sin nombre y prueba']
    if consulta[1] == 3:
        for row in consulta[0]:
            a = row[0]
            b = row[1]
            c = row[2]
            if c and c != "Prueba" and c != "NULL":
                ws4.append([a, b, c])
            else:
                ws7.append(list(row))
    if consulta[1] == 4:
        for row in consulta[0]:
            a = row[0]
            b = row[1]
            c = row[2]
            if c and c != "Prueba" and c != "NULL":
                ws5.append([a, b, c])
            else:
                ws7.append(list(row))

    if consulta[1] == 5:
        for row in consulta[0]:
            a = row[0]
            b = row[1]
            c = row[2]
            if c and c != "Prueba" and c != "NULL":
                ws6.append([a, b, c])
            else:
                ws7.append(list(row))
    wb.save(filename=excelExtensiones)
    wb.close()


# ---------------------------- Tabla dinamica
def tablaDinamica():
    #-------------------- Medellin
    archivoMedellinE = pd.read_excel(
        io=excelExtensiones, sheet_name="Engage - Medellin")
    tablaPM = archivoMedellinE.pivot_table(
        index='Cliente', values='Extension', aggfunc='count', columns='Servidor', margins=True)
    tablaPMR = archivoMedellinE.pivot_table(
        index='Cliente', values='Extension', aggfunc='count', margins=True)
    #-------------------- Peru
    archivoPeruE = pd.read_excel(
        io=excelExtensiones, sheet_name="Engage - Peru")
    tablaPP = archivoPeruE.pivot_table(
        index='Cliente', values='Extension', aggfunc='count',  columns='Servidor', margins=True)
    tablaPPR = archivoPeruE.pivot_table(
        index='Cliente', values='Extension', aggfunc='count', margins=True)
    #-------------------- Bogota
    archivoBogotaE = pd.read_excel(
        io=excelExtensiones, sheet_name="Engage - Bogota")
    tablaPB = archivoBogotaE.pivot_table(
        index='Cliente', values='Extension', aggfunc='count',  columns='Servidor', margins=True)
    tablaPBR = archivoBogotaE.pivot_table(
        index='Cliente', values='Extension', aggfunc='count', margins=True)
    # -------------------- Medellin Up
    archivoMedellinU = pd.read_excel(
        io=excelExtensiones, sheet_name="Uptvity Medellin")
    tablaPMU = archivoMedellinU.pivot_table(
        index='Cliente', values='Extension', aggfunc='count', columns='Servidor', margins=True)
    tablaPMUR = archivoMedellinU.pivot_table(
        index='Cliente', values='Extension', aggfunc='count',  margins=True)
    # -------------------- Bogota Up
    archivoBogotaU = pd.read_excel(
        io=excelExtensiones, sheet_name="Uptvity Bogota")
    tablaPBU = archivoBogotaU.pivot_table(
        index='Cliente', values='Extension', aggfunc='count', columns='Servidor', margins=True)
    tablaPBUR = archivoBogotaU.pivot_table(
        index='Cliente', values='Extension', aggfunc='count', margins=True)
    # -------------------- Despegar Up
    archivoDespegarU = pd.read_excel(
        io=excelExtensiones, sheet_name="Uptivity Despegar")
    tablaPDU = archivoDespegarU.pivot_table(
        index='Cliente', values='Extension', aggfunc='count', columns='Servidor', margins=True)
    tablaPDUR = archivoDespegarU.pivot_table(
        index='Cliente', values='Extension', aggfunc='count', margins=True)

    with pd.ExcelWriter(nombreExcel) as writer:
        tablaPM.to_excel(writer, sheet_name="Medellin Engage")
        tablaPP.to_excel(writer, sheet_name="Perú Engage")
        tablaPB.to_excel(writer, sheet_name="Bogotá Engage")
        tablaPMU.to_excel(writer, sheet_name="Medellin Uptivity")
        tablaPBU.to_excel(writer, sheet_name="Bogotá Uptivity")
        tablaPDU.to_excel(writer, sheet_name="Despegar Uptivity")
        tablaPMR.to_excel(writer, sheet_name="Medellin Engage Conteo")
        tablaPPR.to_excel(writer, sheet_name="Perú Engage Conteo")
        tablaPBR.to_excel(writer, sheet_name="Bogotá Engage Conteo")
        tablaPMUR.to_excel(writer, sheet_name="Medellin Uptivity  Conteo")
        tablaPBUR.to_excel(writer, sheet_name="Bogotá Uptivity Conteo")
        tablaPDUR.to_excel(writer, sheet_name="Despegar Uptivity Conteo")




# ---------------------------- Organizacion del archivo
def organizarExcel():

    # Organizacion del archivo
    wb = load_workbook(nombreExcel)
    BogSheet = wb['Bogotá Engage']
    base = load_workbook(nombreArchivo)
    totalBase = base['Servidores Nice Engage']
    totalBase.column_dimensions['A'].width = 20
    maxFilaB = BogSheet.max_row
    maxColuB = BogSheet.max_column


    # ---------------Bogotá Eng
    for fil in range(1, maxFilaB+1):
        for col in range(1, maxColuB+1):
            # inserto los valores de las demas hojas en la hoja de bogota
            totalBase.cell(row=fil, column=col).value = BogSheet.cell(
                row=fil, column=col).value
            totalBase.cell(row=fil, column=col).border = borde_cuadrado
            # se centra los valores de la tabla
            totalBase.cell(row=fil, column=col).alignment = centrado
            # limpio los valores en las otras Hojas
            # colocamos negrita a la columna
            celdafilB = totalBase.cell(row=fil, column=1)
            # colocamos negrita a la fila
            celdacolB = totalBase.cell(row=1, column=col)
            # -------------------------------------------------------
            celdafilB.font = Font(name='Calibri', size=11, bold=True)
            celdacolB.font = Font(name='Calibri', size=11, bold=True)
            celdafilB.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
            celdacolB.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
    BogSheet.cell(row=1, column=1).value = 'Clientes Bogotá'

# ---------------Medellin Eng
    MedeSheet = wb['Medellin Engage']
    maxFilaM = MedeSheet.max_row
    maxColuM = MedeSheet.max_column
    # for col in range(maxColuP):
    for fil in range(1, maxFilaM+1):
        for col in range(1, maxColuM+1):
            # inserto los valores de las demas hojas en la hoja de bogota
            totalBase.cell(
                row=maxFilaB+fil+3, column=col).value = MedeSheet.cell(row=fil, column=col).value
            # le ponemos el borde a toda la tabla
            totalBase.cell(row=maxFilaB+fil+3,
                           column=col).border = borde_cuadrado
            # se centra los valores de la tabla
            totalBase.cell(row=maxFilaB+fil+3, column=col).alignment = centrado
            # colocamos negrita a la columna
            celdafilM = totalBase.cell(row=maxFilaB+fil+3, column=1)
            # colocamos negrita a la fila
            celdacolM = totalBase.cell(row=maxFilaB+4, column=col)
            # -------------------------------------------------------
            celdafilM.font = Font(name='Calibri', size=11, bold=True)
            celdacolM.font = Font(name='Calibri', size=11, bold=True)
            celdafilM.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
            celdacolM.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
    totalBase.cell(row=maxFilaB+4, column=1).value = 'Clientes Medellin'

# --------------Perú Eng
    PerSheet = wb['Perú Engage']
    maxFilaP = PerSheet.max_row
    maxColuP = PerSheet.max_column
    inicio = maxFilaB+maxFilaM+6
    for fil in range(1, maxFilaP+1):
        for col in range(1, maxColuP+1):
            # inserto los valores de las demas hojas en la hoja de bogota
            totalBase.cell(
                row=inicio+fil, column=col).value = PerSheet.cell(row=fil, column=col).value
            # le ponemos el borde a toda la tabla
            totalBase.cell(row=inicio+fil, column=col).border = borde_cuadrado
            # se centra los valores de la tabla
            totalBase.cell(row=inicio+fil, column=col).alignment = centrado
            # colocamos negrita a la columna
            celdafilP = totalBase.cell(row=inicio+fil, column=1)
            # colocamos negrita a la fila
            celdacolP = totalBase.cell(row=inicio+1, column=col)
            # -------------------------------------------------------
            celdafilP.font = Font(name='Calibri', size=11, bold=True)
            celdacolP.font = Font(name='Calibri', size=11, bold=True)
            celdafilP.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
            celdacolP.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
    BogSheet.cell(row=inicio+1, column=1).value = 'Clientes Perú'

# --------------Bogotá Upt
    BogSheetUp = wb['Bogotá Uptivity']
    maxFilaBU = BogSheetUp.max_row
    maxColuBU = BogSheetUp.max_column
    totalBaseU = base['Servidores Nice Uptivity']
    columnas = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for col in columnas:
        if col == 'A':
            totalBaseU.column_dimensions[col].width = 22
        else:
            totalBaseU.column_dimensions[col].width = 15
    for fil in range(1, maxFilaBU+1):
        for col in range(1, maxColuBU+1):
            # inserto los valores de las demas hojas en la hoja de bogota
            totalBaseU.cell(row=fil, column=col).value = BogSheetUp.cell(
                row=fil, column=col).value
            totalBaseU.cell(row=fil, column=col).border = borde_cuadrado
            # se centra los valores de la tabla
            totalBaseU.cell(row=fil, column=col).alignment = centrado
            # limpio los valores en las otras Hojas
            # colocamos negrita a la columna
            celdafilBU = totalBaseU.cell(row=fil, column=1)
            # colocamos negrita a la fila
            celdacolBU = totalBaseU.cell(row=1, column=col)
            # -------------------------------------------------------
            celdafilBU.font = Font(name='Calibri', size=11, bold=True)
            celdacolBU.font = Font(name='Calibri', size=11, bold=True)
            celdafilBU.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
            celdacolBU.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
    BogSheetUp.cell(row=1, column=1).value = 'Clientes Bogotá'

    # ---------------Medellin Upt
    MedeSheetUp = wb['Medellin Uptivity']
    maxFilaMU = MedeSheetUp.max_row
    maxColuMU = MedeSheetUp.max_column
    # for col in range(maxColuP):
    for fil in range(1, maxFilaMU+1):
        for col in range(1, maxColuMU+1):
            # inserto los valores de las demas hojas en la hoja de bogota
            totalBaseU.cell(
                row=maxFilaBU+fil+3, column=col).value = MedeSheetUp.cell(row=fil, column=col).value
            # le ponemos el borde a toda la tabla
            totalBaseU.cell(row=maxFilaBU+fil+3,
                            column=col).border = borde_cuadrado
            # se centra los valores de la tabla
            totalBaseU.cell(row=maxFilaBU+fil+3,
                            column=col).alignment = centrado
            # colocamos negrita a la columna
            celdafilMU = totalBaseU.cell(row=maxFilaBU+fil+3, column=1)
            # colocamos negrita a la fila
            celdacolMU = totalBaseU.cell(row=maxFilaBU+4, column=col)
            # -------------------------------------------------------
            celdafilMU.font = Font(name='Calibri', size=11, bold=True)
            celdacolMU.font = Font(name='Calibri', size=11, bold=True)
            celdafilMU.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
            celdacolMU.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
    totalBaseU.cell(row=maxFilaBU+4, column=1).value = 'Clientes Medellin'

    # --------------Perú Eng
    DesSheetuP = wb['Despegar Uptivity']
    maxFilaDU = DesSheetuP.max_row
    maxColuDU = DesSheetuP.max_column
    inicioU = maxFilaBU+maxFilaMU+6
    for fil in range(1, maxFilaDU+1):
        for col in range(1, maxColuDU+1):
            # inserto los valores de las demas hojas en la hoja de bogota
            totalBaseU.cell(
                row=inicioU+fil, column=col).value = DesSheetuP.cell(row=fil, column=col).value
            # le ponemos el borde a toda la tabla
            totalBaseU.cell(row=inicioU+fil,
                            column=col).border = borde_cuadrado
            # se centra los valores de la tabla
            totalBaseU.cell(row=inicioU+fil, column=col).alignment = centrado
            # colocamos negrita a la columna
            celdafilP = totalBaseU.cell(row=inicioU+fil, column=1)
            # colocamos negrita a la fila
            celdacolP = totalBaseU.cell(row=inicioU+1, column=col)
            # -------------------------------------------------------
            celdafilP.font = Font(name='Calibri', size=11, bold=True)
            celdacolP.font = Font(name='Calibri', size=11, bold=True)
            celdafilP.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
            celdacolP.fill = PatternFill(
                start_color=colorTabla, end_color=colorTabla, fill_type="solid")
    totalBaseU.cell(row=inicioU+1, column=1).value = 'Clientes Despegar'
    base.save(nombreArchivo)

def combinarTablas():
    wb = load_workbook(nombreExcel)
    MedeSheetEN = wb['Medellin Engage Conteo']
    MedeSheetUP = wb['Medellin Uptivity  Conteo']
    maxFilaME = MedeSheetEN.max_row
    maxFilaMU = MedeSheetUP.max_row
    base = load_workbook(nombreArchivo)
    totalBase = base['Medellin']
    total=[]
    for filEN in range(1,maxFilaME+1):
        if MedeSheetEN.cell(row=filEN, column=1).value != 'All' and MedeSheetEN.cell(row=filEN, column=1).value != 'Cliente':
            total.append((MedeSheetEN.cell(row=filEN, column=1).value).capitalize())
    for filEN in range(1,maxFilaMU+1):
        if MedeSheetUP.cell(row=filEN, column=1).value != 'All' and MedeSheetUP.cell(row=filEN, column=1).value != 'Cliente':
            total.append(((MedeSheetUP.cell(row=filEN, column=1).value).replace(" ","")).capitalize())
    
    salida = pd.unique(total)
    salida.sort()
    for filEN1 in range(len(salida)):
        totalBase.cell(row=filEN1+2, column=1).value = salida[filEN1]
        totalBase.cell(row=filEN1+2,column=1).border = borde_cuadrado
        totalBase.cell(row=filEN1+2, column=1).alignment = centrado
        celdafilP = totalBase.cell(row=filEN1+2, column=1)
        celdacolP = totalBase.cell(row=filEN1+2, column=1)
        for col in range(1,5):
            totalBase.cell(row=filEN1+2,column=col).border = borde_cuadrado
            totalBase.cell(row=filEN1+2, column=col).alignment = centrado
        celdafilP.font = Font(name='Calibri', size=11, bold=True)
        celdacolP.font = Font(name='Calibri', size=11, bold=True)
        celdafilP.fill = PatternFill(start_color=colorTabla, end_color=colorTabla, fill_type="solid")
        celdacolP.fill = PatternFill(start_color=colorTabla, end_color=colorTabla, fill_type="solid")
    
    for filS1 in range(2,len(salida)+2):
        for filS2 in range(2,maxFilaME+1):
            if totalBase.cell(row=filS1, column=1).value == (MedeSheetEN.cell(row=filS2 , column=1).value).capitalize():
                totalBase.cell(row=filS1 , column=2).value = MedeSheetEN.cell(row=filS2 , column=2).value
    
    for filS1 in range(2,len(salida)+2):
        for filS2 in range(2,maxFilaMU+1):
            if totalBase.cell(row=filS1, column=1).value == ((MedeSheetUP.cell(row=filS2, column=1).value).replace(" ","")).capitalize():
                totalBase.cell(row=filS1 , column=3).value = MedeSheetUP.cell(row=filS2 , column=2).value
                


    totalBase.column_dimensions['A'].width = 20


    base.save(nombreArchivo)
    

# ----------- Programa principal consulta Medellin Engage
try:
    cnxn = pyodbc.connect(driver='{SQL Server}',
                          server='JORGERAMIREZ\SQLEXPRESS',
                          database='PRUEBA',
                          trusted_connection='no')
    print("Conexion exitosa")  # en la parte visual colocar algun led
    cursor = cnxn.cursor()  # Cursors represent a database cursor (and map to ODBC HSTMTs), which is used to manage the context of a fetch operation
    crearExcelExt()
    crearExcelFinal()
    med = consultaDbMedEng(cursor)
    guardarExcelEng(med)
    bog = consultaDbBogEng(cursor)
    guardarExcelEng(bog)
    cnxn.close()
except Exception as ex:
    print(ex)

try:
    cnxnu = pyodbc.connect(driver='{SQL Server}',
                           server='JORGERAMIREZ\PRUEBA',
                           database='Uptivity',
                           trusted_connection='no')
    print("Conexion exitosa")  # en la parte visual colocar algun led
    cursor = cnxnu.cursor()
    bog = consultaDbBogUp(cursor)
    guardarExcelUp(bog)
    med = consultaDbMedUp(cursor)
    guardarExcelUp(med)
    des = consultaDbDesUp(cursor)
    guardarExcelUp(des)
    cnxnu.close()
except Exception as ex:
    print(ex)


# ---------------- inicio de la tabla dinamica
tablaDinamica()
# ---------------- Agrupar tablas en una hoja y mejorarlo
organizarExcel()

combinarTablas()
